from __future__ import annotations

import hashlib
import logging
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from unittest.mock import patch

from openpyxl import load_workbook
from pypdf import PdfReader

from ehx_guard.pdf_generator import (
    A5PdfGenerator,
    OfflineOrderLabel,
    PdfGenerationError,
)


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATE = PROJECT_ROOT / "Wologic/System/报交下线单模板.xlsx"


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _label() -> OfflineOrderLabel:
    return OfflineOrderLabel(
        offline_order_no="EHX20260629185500",
        material_code="5664620-CLBK06",
        material_name="主驾座椅背板总成 极夜黑",
        customer_material_code="566462001FA2",
        quantity=6,
        production_time=datetime(2026, 6, 29, 18, 55),
        offline_location="EHX-FG",
    )


class A5PdfGeneratorTest(unittest.TestCase):
    def test_template_copy_replaces_all_tokens_and_preserves_layout(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_text:
            temp_dir = Path(temp_dir_text)
            copy_path = temp_dir / "copy.xlsx"
            generator = A5PdfGenerator(
                TEMPLATE,
                log_path=temp_dir / "pdf.log",
                enable_libreoffice=False,
                enable_excel_com=False,
                barcode_output_dir=temp_dir / "barcodes",
            )
            generator.create_workbook_copy(_label(), copy_path)

            workbook = load_workbook(copy_path)
            try:
                sheet = workbook["标签"]
                text = "\n".join(
                    str(cell.value)
                    for row in sheet.iter_rows()
                    for cell in row
                    if cell.value is not None
                )
                self.assertNotRegex(text, r"\$[A-Za-z0-9_]+\$")
                self.assertIn("5664620-CLBK06", text)
                self.assertIn("2918", text)
                self.assertNotIn("$Reserved1Sub$", text)
                self.assertIn("EHX20260629185500", text)
                self.assertEqual("'标签'!$A$1:$F$12", sheet.print_area)
                self.assertEqual("landscape", sheet.page_setup.orientation)
                self.assertEqual(18, len(sheet.merged_cells.ranges))
                self.assertEqual(0, sheet.page_margins.left)
                self.assertIsNone(sheet["A9"].value)
                self.assertIsNone(sheet["F3"].value)
                self.assertIsNone(sheet["F8"].value)
                self.assertIsNone(sheet["F10"].value)
                self.assertEqual(5, len(sheet._images))
                self.assertEqual(
                    4, len(list((temp_dir / "barcodes").glob("*.png")))
                )
            finally:
                workbook.close()

    def test_company_code_does_not_change_material_barcode(self) -> None:
        label = _label()
        values = label.template_values()
        self.assertEqual("2918", values["Reserved1Sub"])
        self.assertEqual("5664620-CLBK06", values["SapMaterialNo"])
        self.assertNotIn("2918", values["SapMaterialNo"])

    def test_unresolved_placeholder_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_text:
            temp_dir = Path(temp_dir_text)
            altered_template = temp_dir / "template.xlsx"
            workbook = load_workbook(TEMPLATE)
            workbook["标签"]["A13"] = "$UnknownField$"
            workbook.save(altered_template)
            workbook.close()

            generator = A5PdfGenerator(
                altered_template,
                log_path=temp_dir / "pdf.log",
                enable_libreoffice=False,
                enable_excel_com=False,
                barcode_output_dir=temp_dir / "barcodes",
            )
            with self.assertRaisesRegex(
                PdfGenerationError, "模板存在未赋值占位符"
            ):
                generator.create_workbook_copy(_label(), temp_dir / "copy.xlsx")

    def test_libreoffice_failure_logs_and_uses_fallback(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_text:
            temp_dir = Path(temp_dir_text)
            output = temp_dir / "sample.pdf"
            log_path = temp_dir / "pdf.log"
            generator = A5PdfGenerator(
                TEMPLATE,
                log_path=log_path,
                enable_excel_com=False,
                barcode_output_dir=temp_dir / "barcodes",
            )
            generator.soffice_path = Path("/fake/soffice")
            with patch.object(
                generator,
                "_convert_with_libreoffice",
                side_effect=PdfGenerationError("模拟转换错误"),
            ):
                result = generator.generate(_label(), output)
            for handler in generator.logger.handlers:
                handler.flush()

            self.assertTrue(result.is_file())
            self.assertEqual("reportlab_fallback", generator.last_renderer)
            log_text = log_path.read_text(encoding="utf-8")
            self.assertIn("LibreOffice 渲染失败", log_text)
            self.assertIn("模拟转换错误", log_text)
            self.assertIn("fallback PDF renderer", log_text)

    def test_fallback_generates_single_page_a5_landscape(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir_text:
            temp_dir = Path(temp_dir_text)
            output = temp_dir / "sample.pdf"
            generator = A5PdfGenerator(
                TEMPLATE,
                log_path=temp_dir / "pdf.log",
                enable_libreoffice=False,
                enable_excel_com=False,
                barcode_output_dir=temp_dir / "barcodes",
            )
            result = generator.generate(_label(), output)

            reader = PdfReader(str(result))
            self.assertEqual(1, len(reader.pages))
            page = reader.pages[0]
            self.assertGreater(float(page.mediabox.width), float(page.mediabox.height))
            self.assertGreaterEqual(len(page.images), 4)
            self.assertTrue(output.with_suffix(".xlsx").is_file())
            self.assertEqual("reportlab_fallback", generator.last_renderer)

    def test_generation_does_not_modify_original_template(self) -> None:
        before = _sha256(TEMPLATE)
        with tempfile.TemporaryDirectory() as temp_dir_text:
            temp_dir = Path(temp_dir_text)
            generator = A5PdfGenerator(
                TEMPLATE,
                log_path=temp_dir / "pdf.log",
                enable_libreoffice=False,
                enable_excel_com=False,
                barcode_output_dir=temp_dir / "barcodes",
            )
            generator.generate(_label(), temp_dir / "sample.pdf")
        self.assertEqual(before, _sha256(TEMPLATE))

    def test_reject_empty_required_field(self) -> None:
        with self.assertRaises(ValueError):
            OfflineOrderLabel(
                offline_order_no="",
                material_code="5664620-CLBK06",
                material_name="主驾座椅背板总成 极夜黑",
                customer_material_code="566462001FA2",
                quantity=6,
                production_time=datetime(2026, 6, 29, 18, 55),
            )


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    unittest.main()
