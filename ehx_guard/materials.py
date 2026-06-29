"""物料映射导入与条码前缀识别。"""

from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from .database import Database


@dataclass(frozen=True)
class Material:
    material_code: str
    material_name: str
    customer_material_code: str
    box_scan_count: int | None = None


@dataclass(frozen=True)
class MaterialImportResult:
    added: int
    updated: int
    disabled: int


class MaterialRepository:
    """物料号只来自 Excel 或数据库，不在程序中硬编码。"""

    def __init__(
        self,
        database: Database,
        excel_path: str | Path,
        *,
        default_box_scan_count: int = 6,
    ) -> None:
        if int(default_box_scan_count) <= 0:
            raise ValueError("default_box_scan_count 必须大于 0")
        self.database = database
        self.excel_path = Path(excel_path).expanduser().resolve()
        self.default_box_scan_count = int(default_box_scan_count)
        self._materials: list[Material] = []

    def load(self, *, refresh_from_excel: bool = False) -> list[Material]:
        if refresh_from_excel and self.excel_path.is_file():
            self.import_excel()
        rows = self.database.get_enabled_materials()
        if not rows:
            self.import_excel()
            rows = self.database.get_enabled_materials()
        self._materials = [Material(**row) for row in rows]
        self._materials.sort(key=lambda item: len(item.material_code), reverse=True)
        if not self._materials:
            raise RuntimeError("物料配置为空，请导入 EHX物料号匹配.xlsx")
        return list(self._materials)

    def reload(self) -> list[Material]:
        self._materials = []
        return self.load()

    def import_excel(self) -> MaterialImportResult:
        if not self.excel_path.is_file():
            raise FileNotFoundError(f"未找到物料匹配表：{self.excel_path}")
        workbook = load_workbook(
            self.excel_path, read_only=True, data_only=True
        )
        try:
            worksheet = workbook.active
            header_row = self._find_header_row(worksheet)
            materials: list[Material] = []
            seen: set[str] = set()
            for row_number, row in enumerate(
                worksheet.iter_rows(
                    min_row=header_row + 1,
                    min_col=1,
                    max_col=4,
                    values_only=True,
                ),
                start=header_row + 1,
            ):
                code = str(row[0] or "").strip()
                name = str(row[1] or "").strip()
                customer_code = str(row[2] or "").strip()
                box_value = row[3]
                if (
                    not code
                    and not name
                    and not customer_code
                    and (box_value is None or str(box_value).strip() == "")
                ):
                    continue
                if not code or not name or not customer_code:
                    raise ValueError(
                        f"物料表第 {row_number} 行 A/B/C 字段不完整"
                    )
                if code in seen:
                    raise ValueError(
                        f"物料表第 {row_number} 行存在重复物料号：{code}"
                    )
                box_scan_count = self._parse_box_scan_count(
                    box_value, row_number
                )
                seen.add(code)
                materials.append(
                    Material(
                        code,
                        name,
                        customer_code,
                        box_scan_count,
                    )
                )
        finally:
            workbook.close()
        if not materials:
            raise ValueError("物料匹配表没有有效数据")
        result = self.database.sync_materials(materials)
        self._materials = []
        return MaterialImportResult(**result)

    @staticmethod
    def _find_header_row(worksheet: object) -> int:
        for row_number, row in enumerate(
            worksheet.iter_rows(
                min_row=1, max_row=20, min_col=1, max_col=4, values_only=True
            ),
            start=1,
        ):
            normalized = [str(value or "").strip() for value in row]
            if (
                normalized[0]
                in {"佛吉亚料号", "物料号", "物料条码", "物料条码前缀"}
                and ("物料" in normalized[1] or "名称" in normalized[1])
                and "客户" in normalized[2]
                and "数量" in normalized[3]
            ):
                return row_number
        raise ValueError("无法识别物料表表头（预期为 A/B/C/D 四列）")

    def _parse_box_scan_count(self, value: object, row_number: int) -> int:
        if value is None or str(value).strip() == "":
            return self.default_box_scan_count
        if isinstance(value, bool):
            raise ValueError(
                f"物料表第 {row_number} 行每箱数量必须是大于0的整数"
            )
        try:
            numeric = float(value)
        except (TypeError, ValueError) as exc:
            raise ValueError(
                f"物料表第 {row_number} 行每箱数量不是数字：{value}"
            ) from exc
        if not numeric.is_integer() or numeric <= 0:
            raise ValueError(
                f"物料表第 {row_number} 行每箱数量必须是大于0的整数：{value}"
            )
        return int(numeric)

    def identify(self, barcode: str) -> Material | None:
        if not self._materials:
            self.load()
        return next(
            (
                material
                for material in self._materials
                if barcode.startswith(material.material_code)
            ),
            None,
        )

    @staticmethod
    def validate_full_barcode(
        barcode: str, material: Material
    ) -> tuple[bool, str]:
        suffix = barcode[len(material.material_code) :]
        if not re.fullmatch(r"\d{11}", suffix):
            return False, "条码后缀必须为8位日期加3位流水号"
        try:
            datetime.strptime(suffix[:8], "%Y%m%d")
        except ValueError:
            return False, "条码日期格式错误"
        return True, ""
