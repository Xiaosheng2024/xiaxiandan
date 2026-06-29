"""基于客户 Excel 模板生成 A5 横向下线单 PDF。"""

from __future__ import annotations

import gc
import logging
import os
import platform
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from datetime import datetime
from logging.handlers import RotatingFileHandler
from pathlib import Path
from typing import Mapping

from openpyxl import load_workbook
from pypdf import PdfReader
from reportlab.graphics.barcode import code128
from reportlab.lib import colors
from reportlab.lib.pagesizes import A5, landscape
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


class PdfGenerationError(RuntimeError):
    """所有 PDF 渲染器均失败。"""


@dataclass(frozen=True)
class OfflineOrderLabel:
    """模板所需的下线单数据。"""

    offline_order_no: str
    material_code: str
    material_name: str
    customer_material_code: str
    quantity: int
    production_time: datetime
    supplier_name: str = "Faurecia"
    customer_name: str = "SEBANGO"
    offline_location: str = "---"
    reserved1_sub: str = "2918"

    def __post_init__(self) -> None:
        required = {
            "offline_order_no": self.offline_order_no,
            "material_code": self.material_code,
            "material_name": self.material_name,
            "customer_material_code": self.customer_material_code,
            "reserved1_sub": self.reserved1_sub,
        }
        missing = [name for name, value in required.items() if not str(value).strip()]
        if missing:
            raise ValueError(f"下线单必填字段为空：{', '.join(missing)}")
        if self.quantity <= 0:
            raise ValueError("quantity 必须大于 0")

    def template_values(self) -> Mapping[str, str]:
        return {
            "SupplierName": self.supplier_name,
            "PartName": self.material_name,
            "CustomerName": self.customer_name,
            "CustomerPartNo": self.customer_material_code,
            "UpdatedAt": self.production_time.strftime("P%Y%m%d %H:%M"),
            "Reserved1Sub": self.reserved1_sub,
            "SapMaterialNo": self.material_code,
            "BoxQty": str(self.quantity),
            "OfflineLocation": self.offline_location,
            "Batch": self.offline_order_no,
        }


class A5PdfGenerator:
    """从模板副本生成单页 A5 横向 PDF，绝不写入原始模板。"""

    def __init__(
        self,
        template_path: str | Path,
        *,
        soffice_path: str | Path | None = None,
        temp_root: str | Path | None = None,
        log_path: str | Path = "logs/pdf_generation.log",
        enable_libreoffice: bool = True,
        enable_excel_com: bool = False,
        enable_office_pdf_on_mac: bool = False,
    ) -> None:
        self.template_path = Path(template_path).expanduser().resolve()
        if not self.template_path.is_file():
            raise FileNotFoundError(f"未找到下线单模板：{self.template_path}")
        office_allowed = enable_libreoffice and (
            platform.system() != "Darwin" or enable_office_pdf_on_mac
        )
        self.soffice_path = find_soffice(soffice_path) if office_allowed else None
        self.office_pdf_skipped_on_mac = (
            enable_libreoffice
            and platform.system() == "Darwin"
            and not enable_office_pdf_on_mac
        )
        self.enable_excel_com = enable_excel_com
        self.temp_root = (
            Path(temp_root).expanduser().resolve() if temp_root else None
        )
        self.logger = _create_file_logger(Path(log_path).expanduser().resolve())
        self.last_renderer: str | None = None

    def generate(
        self,
        label: OfflineOrderLabel,
        output_path: str | Path,
    ) -> Path:
        """按当前操作系统选择渲染器并生成 PDF。"""

        target = Path(output_path).expanduser().resolve()
        if target.suffix.lower() != ".pdf":
            raise ValueError("output_path 必须以 .pdf 结尾")
        target.parent.mkdir(parents=True, exist_ok=True)
        if self.temp_root:
            self.temp_root.mkdir(parents=True, exist_ok=True)

        with tempfile.TemporaryDirectory(
            prefix="ehx_pdf_", dir=self.temp_root
        ) as temp_dir_text:
            temp_dir = Path(temp_dir_text)
            workbook_path = temp_dir / "offline_order.xlsx"
            self.create_workbook_copy(label, workbook_path)
            generated_pdf = self._render_with_fallbacks(
                label, workbook_path, temp_dir
            )
            _validate_a5_landscape_pdf(generated_pdf)
            _atomic_copy(generated_pdf, target)

        self.logger.info(
            "PDF 生成成功 renderer=%s order=%s output=%s",
            self.last_renderer,
            label.offline_order_no,
            target,
        )
        return target

    def create_workbook_copy(
        self,
        label: OfflineOrderLabel,
        destination: str | Path,
    ) -> Path:
        """创建已替换占位符的模板副本，保留原模板不变。"""

        destination_path = Path(destination).expanduser().resolve()
        destination_path.parent.mkdir(parents=True, exist_ok=True)
        workbook = load_workbook(self.template_path)
        values = label.template_values()

        replaced: set[str] = set()
        unresolved: set[str] = set()
        token_pattern = re.compile(r"\$([A-Za-z0-9_]+)\$")
        try:
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows():
                    for cell in row:
                        if not isinstance(cell.value, str):
                            continue
                        text = cell.value
                        for token in token_pattern.findall(text):
                            if token in values:
                                text = text.replace(f"${token}$", values[token])
                                replaced.add(token)
                        cell.value = text
                        unresolved.update(token_pattern.findall(text))

            missing_from_template = set(values) - replaced
            if missing_from_template:
                raise PdfGenerationError(
                    "模板缺少占位符：" + ", ".join(sorted(missing_from_template))
                )
            if unresolved:
                raise PdfGenerationError(
                    "模板存在未赋值占位符：" + ", ".join(sorted(unresolved))
                )
            workbook.save(destination_path)
        finally:
            workbook.close()

        self.logger.info("模板副本生成成功 path=%s", destination_path)
        return destination_path

    def _render_with_fallbacks(
        self,
        label: OfflineOrderLabel,
        workbook_path: Path,
        temp_dir: Path,
    ) -> Path:
        failures: list[str] = []

        # Windows 正式环境优先 LibreOffice；macOS 默认在构造器中禁用。
        if self.soffice_path:
            try:
                output = temp_dir / "libreoffice.pdf"
                self._convert_with_libreoffice(workbook_path, output, temp_dir)
                _validate_a5_landscape_pdf(output)
                self.last_renderer = "libreoffice"
                return output
            except Exception as exc:
                detail = f"LibreOffice 渲染失败：{exc}"
                failures.append(detail)
                self.logger.exception(detail)
        else:
            if self.office_pdf_skipped_on_mac:
                self.logger.info(
                    "macOS 默认跳过 LibreOffice；"
                    "enable_office_pdf_on_mac=false"
                )
            else:
                self.logger.warning("LibreOffice 不可用，跳过 LibreOffice 渲染")

        # 兼容保留：仅显式启用时在 Windows 尝试 Excel COM。
        if self.enable_excel_com and os.name == "nt":
            excel_available, excel_detail = is_excel_com_available()
            if excel_available:
                try:
                    output = temp_dir / "excel_com.pdf"
                    self._convert_with_excel_com(workbook_path, output)
                    _validate_a5_landscape_pdf(output)
                    self.last_renderer = "excel_com"
                    return output
                except Exception as exc:
                    detail = f"Excel COM 渲染失败：{exc}"
                    failures.append(detail)
                    self.logger.exception(detail)
            else:
                self.logger.warning("Excel COM 不可用：%s", excel_detail)

        try:
            output = temp_dir / "reportlab_fallback.pdf"
            self.logger.warning(
                "使用 fallback PDF renderer；此前失败：%s",
                " | ".join(failures) if failures else "首选渲染器不可用",
            )
            self._render_with_reportlab(label, output)
            self.last_renderer = "reportlab_fallback"
            return output
        except Exception as exc:
            failures.append(f"ReportLab fallback 渲染失败：{exc}")
            self.logger.exception("ReportLab fallback 渲染失败")
            raise PdfGenerationError("；".join(failures)) from exc

    def _convert_with_libreoffice(
        self,
        workbook_path: Path,
        destination: Path,
        temp_dir: Path,
    ) -> None:
        if not self.soffice_path or not self.soffice_path.is_file():
            raise PdfGenerationError("LibreOffice 命令不存在或不可访问")

        conversion_dir = temp_dir / "libreoffice_output"
        conversion_dir.mkdir()
        profile_dir = temp_dir / "libreoffice_profile"
        command = [
            str(self.soffice_path),
            "--headless",
            f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
            "--convert-to",
            "pdf",
            "--outdir",
            str(conversion_dir),
            str(workbook_path),
        ]
        startup_info, creation_flags = _hidden_process_options()
        try:
            result = subprocess.run(
                command,
                capture_output=True,
                text=True,
                timeout=60,
                check=False,
                startupinfo=startup_info,
                creationflags=creation_flags,
            )
        except (OSError, subprocess.TimeoutExpired) as exc:
            raise PdfGenerationError(f"调用 LibreOffice 失败：{exc}") from exc

        self.logger.info(
            "LibreOffice command=%s returncode=%s stdout=%s stderr=%s",
            command,
            result.returncode,
            result.stdout.strip(),
            result.stderr.strip(),
        )
        if result.returncode != 0:
            detail = (result.stderr or result.stdout or "无错误信息").strip()
            raise PdfGenerationError(
                f"退出码 {result.returncode}，详情：{detail}"
            )
        converted = conversion_dir / f"{workbook_path.stem}.pdf"
        if not converted.is_file():
            raise PdfGenerationError("命令成功返回，但没有生成 PDF 文件")
        shutil.move(converted, destination)

    def _convert_with_excel_com(
        self,
        workbook_path: Path,
        destination: Path,
    ) -> None:
        if os.name != "nt":
            raise PdfGenerationError("Excel COM 仅支持 Windows")
        try:
            import pythoncom
            import win32com.client
        except ImportError as exc:
            raise PdfGenerationError("未安装 pywin32") from exc

        excel = None
        workbook = None
        pythoncom.CoInitialize()
        try:
            excel = win32com.client.DispatchEx("Excel.Application")
            excel.Visible = False
            excel.DisplayAlerts = False
            workbook = excel.Workbooks.Open(
                str(workbook_path), UpdateLinks=0, ReadOnly=True
            )
            # 0 = xlTypePDF。Excel 会沿用模板打印区域、纸张方向和分页设置。
            workbook.ExportAsFixedFormat(
                Type=0,
                Filename=str(destination),
                Quality=0,
                IncludeDocProperties=True,
                IgnorePrintAreas=False,
                OpenAfterPublish=False,
            )
        finally:
            if workbook is not None:
                try:
                    workbook.Close(SaveChanges=False)
                except Exception:
                    self.logger.exception("关闭 Excel 工作簿失败")
            if excel is not None:
                try:
                    excel.Quit()
                except Exception:
                    self.logger.exception("退出 Excel COM 进程失败")
            workbook = None
            excel = None
            gc.collect()
            pythoncom.CoUninitialize()
        if not destination.is_file():
            raise PdfGenerationError("Excel COM 未生成 PDF 文件")

    def _render_with_reportlab(
        self,
        label: OfflineOrderLabel,
        destination: Path,
    ) -> None:
        _register_chinese_font()
        page_width, page_height = landscape(A5)
        pdf = canvas.Canvas(
            str(destination),
            pagesize=(page_width, page_height),
            pageCompression=1,
        )
        pdf.setTitle(f"EHX 下线单 {label.offline_order_no}")
        pdf.setAuthor(label.supplier_name)

        margin = 12
        bottom = 12
        top = page_height - 12
        left_width = 145
        center_width = 275
        x0 = margin
        x1 = x0 + left_width
        x2 = x1 + center_width
        x3 = page_width - margin

        pdf.setStrokeColor(colors.black)
        pdf.setLineWidth(0.8)
        pdf.rect(x0, bottom, x3 - x0, top - bottom)
        pdf.line(x1, bottom, x1, top)
        pdf.line(x2, bottom, x2, top)

        # 左区：供应商、生产时间、供应商物料号和库位。
        y_supplier = top - 76
        y_time = top - 148
        y_material = top - 250
        pdf.line(x0, y_supplier, x1, y_supplier)
        pdf.line(x0, y_time, x1, y_time)
        pdf.line(x0, y_material, x1, y_material)
        _draw_text(pdf, label.supplier_name, x0 + 8, top - 30, 14)
        _draw_label(pdf, "生产时间", x0 + 6, y_supplier - 16)
        _draw_text(
            pdf,
            label.production_time.strftime("P%Y%m%d %H:%M"),
            x0 + 6,
            y_supplier - 43,
            13,
        )
        _draw_label(pdf, "供应商零件号", x0 + 6, y_time - 16)
        _draw_text(pdf, label.material_code, x0 + 6, y_time - 44, 13)
        _draw_code128(
            pdf, label.material_code, x0 + 7, y_time - 91, left_width - 14, 34
        )
        _draw_label(pdf, "产品类型", x0 + 6, y_material - 18)
        _draw_text(pdf, "FG", x0 + 18, bottom + 28, 22)
        _draw_label(pdf, "下线库位", x0 + 72, y_material - 18)
        _draw_text(pdf, label.offline_location, x0 + 68, bottom + 28, 15)

        # 中区：物料描述、客户/项目名和醒目的物料短码。
        y_description = top - 102
        pdf.line(x1, y_description, x2, y_description)
        _draw_label(pdf, "零件描述", x1 + 6, top - 16)
        _draw_centered_text(
            pdf, label.material_name, x1, x2, top - 54, 17, max_width=260
        )
        pdf.setFillColor(colors.black)
        pdf.rect(x1, y_description - 29, center_width, 29, fill=1, stroke=0)
        pdf.setFillColor(colors.white)
        _draw_centered_text(
            pdf, label.customer_name, x1, x2, y_description - 20, 11
        )
        pdf.setFillColor(colors.black)
        _draw_centered_text(
            pdf,
            label.reserved1_sub,
            x1,
            x2,
            y_description - 135,
            62,
            max_width=255,
        )
        _draw_centered_text(
            pdf, "MADE IN CHINA", x1, x2, bottom + 14, 10
        )

        # 右区：客户料号、数量和批次/下线单号，均附可扫描 Code 128。
        y_customer = top - 108
        y_quantity = top - 230
        pdf.line(x2, y_customer, x3, y_customer)
        pdf.line(x2, y_quantity, x3, y_quantity)
        _draw_label(pdf, "客户零件号", x2 + 90, top - 16)
        _draw_centered_text(
            pdf, label.customer_material_code, x2, x3, top - 49, 16
        )
        _draw_code128(
            pdf,
            label.customer_material_code,
            x2 + 12,
            y_customer + 12,
            x3 - x2 - 24,
            30,
        )
        _draw_label(pdf, "数量", x3 - 34, y_customer - 17)
        _draw_centered_text(
            pdf, str(label.quantity), x2, x3, y_customer - 55, 31
        )
        _draw_code128(
            pdf, str(label.quantity), x2 + 43, y_quantity + 12, x3 - x2 - 86, 25
        )
        _draw_label(pdf, "批次号 / 下线单号", x2 + 64, y_quantity - 18)
        _draw_centered_text(
            pdf, label.offline_order_no, x2, x3, y_quantity - 53, 13
        )
        _draw_code128(
            pdf,
            label.offline_order_no,
            x2 + 12,
            bottom + 22,
            x3 - x2 - 24,
            34,
        )

        pdf.showPage()
        pdf.save()


def find_soffice(explicit_path: str | Path | None = None) -> Path | None:
    """查找 LibreOffice 命令；找不到时返回 None，不阻止 fallback。"""

    if explicit_path:
        candidate = Path(explicit_path).expanduser().resolve()
        return candidate if candidate.is_file() else None
    discovered = shutil.which("soffice") or shutil.which("libreoffice")
    if discovered:
        return Path(discovered).resolve()
    for candidate in (
        Path(r"C:\Program Files\LibreOffice\program\soffice.exe"),
        Path(r"C:\Program Files (x86)\LibreOffice\program\soffice.exe"),
        Path("/Applications/LibreOffice.app/Contents/MacOS/soffice"),
    ):
        if candidate.is_file():
            return candidate.resolve()
    return None


def is_excel_com_available() -> tuple[bool, str]:
    """只在 Windows 检查 pywin32 和 Excel COM 注册信息。"""

    if os.name != "nt":
        return False, "非 Windows 系统"
    try:
        import winreg
        import win32com.client  # noqa: F401
    except ImportError as exc:
        return False, f"依赖不可用：{exc}"
    try:
        key = winreg.OpenKey(winreg.HKEY_CLASSES_ROOT, r"Excel.Application\CLSID")
        winreg.CloseKey(key)
    except OSError:
        return False, "未检测到 Microsoft Excel COM 注册信息"
    return True, "Microsoft Excel COM 可用"


def _create_file_logger(log_path: Path) -> logging.Logger:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger(f"ehx_guard.pdf.{log_path}")
    logger.setLevel(logging.INFO)
    logger.propagate = False
    if not logger.handlers:
        handler = RotatingFileHandler(
            log_path,
            maxBytes=2 * 1024 * 1024,
            backupCount=5,
            encoding="utf-8",
        )
        handler.setFormatter(
            logging.Formatter(
                "%(asctime)s %(levelname)s %(name)s %(message)s"
            )
        )
        logger.addHandler(handler)
    return logger


def _hidden_process_options() -> tuple[object | None, int]:
    if os.name != "nt":
        return None, 0
    startup_info = subprocess.STARTUPINFO()
    startup_info.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    return startup_info, subprocess.CREATE_NO_WINDOW


def _atomic_copy(source: Path, target: Path) -> None:
    staged_target = target.with_name(f".{target.name}.tmp")
    try:
        shutil.copyfile(source, staged_target)
        os.replace(staged_target, target)
    finally:
        staged_target.unlink(missing_ok=True)


def _validate_a5_landscape_pdf(path: Path) -> None:
    reader = PdfReader(str(path))
    if len(reader.pages) != 1:
        raise PdfGenerationError(
            f"下线单必须为单页，实际生成 {len(reader.pages)} 页"
        )
    page = reader.pages[0]
    width_mm = float(page.mediabox.width) * 25.4 / 72
    height_mm = float(page.mediabox.height) * 25.4 / 72
    if width_mm < height_mm:
        raise PdfGenerationError("生成的 PDF 不是横向页面")
    if abs(width_mm - 210) > 3 or abs(height_mm - 148) > 3:
        raise PdfGenerationError(
            f"PDF 页面不是 A5：实际约 {width_mm:.1f} × {height_mm:.1f} mm"
        )


_FALLBACK_FONT_NAME = "STSong-Light"


def _register_chinese_font() -> None:
    global _FALLBACK_FONT_NAME
    if "EHXChinese" in pdfmetrics.getRegisteredFontNames():
        _FALLBACK_FONT_NAME = "EHXChinese"
        return
    font_candidates = (
        Path(r"C:\Windows\Fonts\msyh.ttc"),
        Path(r"C:\Windows\Fonts\simhei.ttf"),
        Path(r"C:\Windows\Fonts\simsun.ttc"),
        Path("/Library/Fonts/Arial Unicode.ttf"),
        Path("/System/Library/Fonts/STHeiti Light.ttc"),
        Path("/usr/share/fonts/opentype/noto/NotoSansCJK-Regular.ttc"),
    )
    for font_path in font_candidates:
        if not font_path.is_file():
            continue
        try:
            pdfmetrics.registerFont(
                TTFont("EHXChinese", str(font_path), subfontIndex=0)
            )
            _FALLBACK_FONT_NAME = "EHXChinese"
            return
        except Exception:
            continue
    if "STSong-Light" not in pdfmetrics.getRegisteredFontNames():
        pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    _FALLBACK_FONT_NAME = "STSong-Light"


def _draw_label(
    pdf: canvas.Canvas, text: str, x: float, y: float, size: float = 8
) -> None:
    _draw_text(pdf, text, x, y, size)


def _draw_text(
    pdf: canvas.Canvas, text: str, x: float, y: float, size: float
) -> None:
    pdf.setFont(_FALLBACK_FONT_NAME, size)
    pdf.drawString(x, y, str(text))


def _draw_centered_text(
    pdf: canvas.Canvas,
    text: str,
    x1: float,
    x2: float,
    y: float,
    size: float,
    *,
    max_width: float | None = None,
) -> None:
    text = str(text)
    width_limit = max_width or (x2 - x1 - 8)
    actual_size = size
    while (
        actual_size > 7
        and pdfmetrics.stringWidth(text, _FALLBACK_FONT_NAME, actual_size)
        > width_limit
    ):
        actual_size -= 0.5
    pdf.setFont(_FALLBACK_FONT_NAME, actual_size)
    pdf.drawCentredString((x1 + x2) / 2, y, text)


def _draw_code128(
    pdf: canvas.Canvas,
    value: str,
    x: float,
    y: float,
    max_width: float,
    height: float,
) -> None:
    barcode = code128.Code128(
        str(value), barHeight=height, barWidth=0.7, humanReadable=False
    )
    if barcode.width > max_width:
        scale = max_width / barcode.width
        pdf.saveState()
        pdf.translate(x, y)
        pdf.scale(scale, 1)
        barcode.drawOn(pdf, 0, 0)
        pdf.restoreState()
    else:
        barcode.drawOn(pdf, x + (max_width - barcode.width) / 2, y)
